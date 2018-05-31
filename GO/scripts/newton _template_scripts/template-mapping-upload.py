from openpyxl import load_workbook
from redis import StrictRedis
from argparse import ArgumentParser
import simplejson as json


class TemplateMapping(object):

    """
    Template mapping via an excel documents.
    """

    def __init__(self, filename, redis_host):
        self.column_list = [
            'OS', 'DB', 'AV', 'OSR', 'MBU', 'Master ID', 'Device ID']
        self.column_letter_list = []
        self.format = 'Master ID.OS.DB.AV.OSR.MBU'
        self.key = 'newton:mapping:'
        self.wb = load_workbook(filename, use_iterators=True, data_only=True)
        self.redis = StrictRedis(redis_host)

    def _lookup_column_by_name(self, ws_name, product_type):
        for worksheet in self.data['worksheets']:
            if worksheet['name'] == ws_name:
                for column in worksheet['columns']:
                    if column['value'] == product_type:
                        return column['column']

    def _return_columns_from_worksheet(self):
        self.data = {}
        self.data['worksheets'] = []

        for ws in self.wb:
            ws_data = {}
            ws_data['name'] = ws.title
            ws_data['columns'] = []
            for row in ws.iter_rows():
                for cell in row:
                    if (cell.row == 1 and
                            cell.value in self.column_list):
                        if cell.column not in self.column_letter_list:
                            cell_data = {}
                            cell_data['column'] = cell.column
                            cell_data['value'] = cell.value
                            ws_data['columns'].append(cell_data)
            self.data['worksheets'].append(ws_data)
        print self.data['worksheets']

    def _return_templates_by_worksheet(self):
        self.templates = {}
        self.templates['worksheets'] = []

        for worksheet in self.data['worksheets']:
            temp_work = {}
            temp_work['name'] = worksheet['name']
            temp_work['templates'] = []
            ws = self.wb.get_sheet_by_name(worksheet['name'])
            for row in ws.iter_rows():
                key_dict = {}
                for cell in row:
                    for sub_item in self.column_list:
                        column = self._lookup_column_by_name(
                            worksheet['name'], sub_item)
                        if (cell.column == column and
                                cell.value is not None and
                                cell.row != 1):
                            try:
                                data = cell.value.split(
                                    '-')[0].replace(' ', '')
                            except Exception:
                                data = str(int(cell.value))
                            try:
                                desc = cell.value.split(
                                    '-')[1]
                            except Exception:
                                desc = ''
                            if data != '':
                                key_dict[sub_item] = {}
                                key_dict[sub_item]['id'] = data
                                key_dict[sub_item]['desc'] = desc
                temp_work['templates'].append(key_dict)
            self.templates['worksheets'].append(temp_work)
        print "Template", self.templates['worksheets']

    def save_templates_to_redis(self):
        print "Save temp into redis"
        self._return_columns_from_worksheet()
        self._return_templates_by_worksheet()
        for worksheet in self.templates['worksheets']:
            for template in worksheet['templates']:
                temp = ''
                desc = ''
                if template != {}:
                    if template.get('Device ID', None):
                        clone_source = template.get('Device ID')
                    for sub_item in self.format.split('.'):
                        if template.get(sub_item, None):
                            temp = template.get(sub_item)['id'] + '.' + temp
                    for sub_item in reversed(self.format.split('.')):
                            try:
                                desc = template.get(sub_item)[
                                    'desc'] + '\n' + desc
                            except TypeError:
                                pass
                    k = temp.rfind(".")
                    new_string = temp[:k] + '' + temp[k + 1:]
                    key = 'newton:mapping:' + new_string
                    value = {'clone_source': clone_source[
                        'id'], 'name': new_string,
                        'description': 'HardwareTemplate:%s %s' %
                        (clone_source['id'], desc)}

                    self.redis.set(key, json.dumps(value))


if __name__ == '__main__':
    parser = ArgumentParser(description='Template Mapping Importer')
    parser.add_argument('-r', '--redis', help='redis host')
    parser.add_argument('-f', '--filename', help='excel filename')
    args = parser.parse_args()
    mapping = TemplateMapping(
        filename=args.filename,
        redis_host=args.redis
        )
    mapping.save_templates_to_redis()
